"""
Invoice Generator
-----------------
Reads an import Excel file (invoice + bill_audit sheets) and outputs,
per client: one PDF invoice (Service Invoice-Charges) and one Excel
file (Service Invoice-Charges + Transaction Record).
"""

import os
import sys
import shutil
import threading
import traceback
from datetime import datetime

import pythoncom
import win32com.client
import openpyxl
from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resource_path(relative_path: str) -> str:
    """Return absolute path to a resource, works both in dev and PyInstaller."""
    try:
        base = sys._MEIPASS  # type: ignore[attr-defined]
    except AttributeError:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)



# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------

def read_import(path: str):
    """
    Returns:
        invoice_data : {client_code: {col_header: value, ...}}
        bill_data    : {client_code: [{col_header: value, ...}, ...]}
    """
    wb = load_workbook(path, data_only=True)

    # ── invoice sheet ──────────────────────────────────────────────────────
    ws_inv = wb["invoice"]
    headers_inv = [c.value for c in next(ws_inv.iter_rows(min_row=1, max_row=1))]
    invoice_data: dict = {}
    for row in ws_inv.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        client_code = str(row[0]).strip()
        invoice_data[client_code] = dict(zip(headers_inv, row))

    # ── bill_audit sheet ───────────────────────────────────────────────────
    ws_bill = wb["bill_audit"]
    headers_bill = [c.value for c in next(ws_bill.iter_rows(min_row=1, max_row=1))]
    bill_data: dict = {}
    for row in ws_bill.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers_bill, row))
        client_code = str(row_dict.get("客戶代碼", "")).strip()
        if client_code:
            bill_data.setdefault(client_code, []).append(row_dict)

    return invoice_data, bill_data


def fill_invoice_sheet(ws, client_code: str, info: dict, date_str: str):
    """Write client data into the Service Invoice-Charges worksheet."""
    safe_date = date_str.replace("/", "").replace("-", "")
    invoice_no = f"{client_code}{safe_date}"

    price_raw = info.get("未稅價格") or 0
    try:
        price = float(price_raw)
    except (TypeError, ValueError):
        price = 0.0

    ws["G2"] = invoice_no          # Invoice No
    ws["G3"] = date_str            # Date
    ws["G4"] = client_code         # Customer ID
    ws["C9"] = info.get("客戶姓名") or ""
    ws["C10"] = info.get("公司名稱") or ""
    ws["C12"] = info.get("客戶地址") or ""
    ws["C14"] = info.get("客戶電話") or ""

    # Set price as numeric so the template's SUM/ROUNDUP formulas still work.
    # Apply a $ number format so it displays as "$1,715" in Excel / PDF.
    dollar_fmt = '"$"#,##0.##'
    ws["E17"] = price
    ws["E17"].number_format = dollar_fmt   # Price (excl. VAT)
    ws["G17"] = price
    ws["G17"].number_format = dollar_fmt   # Line Total
    # G27 = SUM(G17:G26), G28 = ROUNDUP(…*5%,0), G29 = SUM(G27:G28)
    # — these formulas are already in the template; do NOT overwrite them.


def fill_transaction_sheet(ws, transactions: list):
    """Clear data rows then fill Transaction Record from bill_audit rows."""
    # Clear all data rows while keeping header
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # bill_audit column  →  Transaction Record column (1-based)
    # A=1 ClientCode, B=2 TrackingNumber, C=3 Destination, D=4 Currency,
    # E=5 ChargeWeight, F=6 Freight, G=7 Tracking Fee, H=8 AdditionalFee,
    # I=9 TotalCost, J=10 ArrivalDate
    for i, txn in enumerate(transactions):
        r = i + 2

        arrival = txn.get("到貨時間")
        if arrival and hasattr(arrival, "strftime"):
            arrival = arrival.strftime("%Y/%m/%d")
        elif arrival:
            arrival = str(arrival).split(" ")[0]

        ws.cell(r, 1).value = txn.get("客戶代碼", "")
        ws.cell(r, 2).value = txn.get("業務單號", "")
        ws.cell(r, 3).value = txn.get("目的地", "")
        ws.cell(r, 4).value = "TWD"
        ws.cell(r, 5).value = txn.get("計費重", "")
        ws.cell(r, 6).value = txn.get("A價_運費", "")
        ws.cell(r, 7).value = txn.get("A價_掛號", "")
        ws.cell(r, 8).value = txn.get("A價_附加費(TWD)", "")
        ws.cell(r, 9).value = txn.get("A價總額(TWD)", "")
        ws.cell(r, 10).value = arrival


def export_first_sheet_as_pdf(excel_abs_path: str, pdf_abs_path: str):
    """Use Excel COM automation to export the first worksheet as PDF."""
    pythoncom.CoInitialize()
    excel_app = None
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        wb = excel_app.Workbooks.Open(excel_abs_path)
        try:
            ws = wb.Worksheets(1)
            ws.Select()
            ws.ExportAsFixedFormat(
                Type=0,                    # xlTypePDF
                Filename=pdf_abs_path,
                Quality=0,                 # xlQualityStandard
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False,
            )
        finally:
            wb.Close(SaveChanges=False)
    finally:
        if excel_app is not None:
            excel_app.Quit()
        pythoncom.CoUninitialize()


def generate_client_files(
    client_code: str,
    info: dict,
    transactions: list,
    output_dir: str,
    invoice_date: str,
    log_fn,
):
    template_path = resource_path("Invoice_Sample.xlsx")
    safe_date = invoice_date.replace("/", "").replace("-", "")

    excel_name = f"{client_code}_{safe_date}_invoice.xlsx"
    pdf_name   = f"{client_code}_{safe_date}_invoice.pdf"
    excel_path = os.path.join(output_dir, excel_name)
    pdf_path   = os.path.join(output_dir, pdf_name)

    # Copy template and fill
    shutil.copy(template_path, excel_path)
    wb = load_workbook(excel_path)

    fill_invoice_sheet(wb["Service Invoice-Charges"], client_code, info, invoice_date)
    fill_transaction_sheet(wb["Transaction Record"], transactions)

    wb.save(excel_path)
    log_fn(f"  [Excel] {excel_name}")

    # Export first sheet as PDF
    export_first_sheet_as_pdf(os.path.abspath(excel_path), os.path.abspath(pdf_path))
    log_fn(f"  [PDF]   {pdf_name}")


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("Invoice Generator")
        root.geometry("680x560")
        root.resizable(False, False)
        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")

        main = ttk.Frame(self.root, padding=18)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Invoice Generator", font=("Arial", 15, "bold")).grid(
            row=0, column=0, columnspan=3, pady=(0, 14), sticky="w"
        )

        # Import file
        ttk.Label(main, text="Import File:").grid(row=1, column=0, sticky="w", pady=6)
        self.import_var = tk.StringVar()
        ttk.Entry(main, textvariable=self.import_var, width=50).grid(
            row=1, column=1, padx=6, pady=6
        )
        ttk.Button(main, text="Browse…", command=self._browse_import).grid(row=1, column=2)

        # Output folder
        ttk.Label(main, text="Output Folder:").grid(row=2, column=0, sticky="w", pady=6)
        self.output_var = tk.StringVar()
        ttk.Entry(main, textvariable=self.output_var, width=50).grid(
            row=2, column=1, padx=6, pady=6
        )
        ttk.Button(main, text="Browse…", command=self._browse_output).grid(row=2, column=2)

        # Invoice date
        ttk.Label(main, text="Invoice Date:").grid(row=3, column=0, sticky="w", pady=6)
        self.date_var = tk.StringVar(value=datetime.today().strftime("%Y/%m/%d"))
        date_frame = ttk.Frame(main)
        date_frame.grid(row=3, column=1, sticky="w", padx=6)
        ttk.Entry(date_frame, textvariable=self.date_var, width=18).pack(side="left")
        ttk.Label(date_frame, text="  (YYYY/MM/DD)", foreground="#666").pack(side="left")

        # Buttons
        btn_frame = ttk.Frame(main)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=14)
        self.gen_btn = ttk.Button(
            btn_frame, text="Generate Invoices", command=self._start_generate, width=22
        )
        self.gen_btn.pack(side="left", padx=6)
        self.open_btn = ttk.Button(
            btn_frame, text="Open Output Folder", command=self._open_output, width=20, state="disabled"
        )
        self.open_btn.pack(side="left", padx=6)

        # Progress bar
        self.progress = ttk.Progressbar(main, mode="indeterminate", length=580)
        self.progress.grid(row=5, column=0, columnspan=3, pady=(0, 6))

        # Status label
        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(main, textvariable=self.status_var, foreground="#444").grid(
            row=6, column=0, columnspan=3, sticky="w"
        )

        # Log
        ttk.Label(main, text="Log:").grid(row=7, column=0, sticky="w", pady=(8, 2))
        self.log_box = scrolledtext.ScrolledText(
            main, height=13, width=76, font=("Courier New", 9), state="disabled"
        )
        self.log_box.grid(row=8, column=0, columnspan=3, pady=4)

    # ── Events ────────────────────────────────────────────────────────────

    def _browse_import(self):
        p = filedialog.askopenfilename(
            title="Select Import File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if p:
            self.import_var.set(p)

    def _browse_output(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            self.output_var.set(p)

    def _open_output(self):
        path = self.output_var.get().strip()
        if path and os.path.isdir(path):
            os.startfile(path)

    # ── Logging ───────────────────────────────────────────────────────────

    def log(self, msg: str):
        self.log_box.configure(state="normal")
        self.log_box.insert(tk.END, msg + "\n")
        self.log_box.see(tk.END)
        self.log_box.configure(state="disabled")
        self.root.update_idletasks()

    def set_status(self, msg: str):
        self.status_var.set(msg)
        self.root.update_idletasks()

    # ── Generate ──────────────────────────────────────────────────────────

    def _start_generate(self):
        import_path = self.import_var.get().strip()
        output_path = self.output_var.get().strip()
        invoice_date = self.date_var.get().strip()

        if not import_path:
            messagebox.showerror("Error", "Please select an import file.")
            return
        if not os.path.isfile(import_path):
            messagebox.showerror("Error", f"Import file not found:\n{import_path}")
            return
        if not output_path:
            messagebox.showerror("Error", "Please select an output folder.")
            return
        if not os.path.isdir(output_path):
            messagebox.showerror("Error", f"Output folder not found:\n{output_path}")
            return
        if not invoice_date:
            messagebox.showerror("Error", "Please enter an invoice date.")
            return

        # Clear log
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", tk.END)
        self.log_box.configure(state="disabled")

        self.gen_btn.state(["disabled"])
        self.open_btn.state(["disabled"])
        self.progress.start(10)
        self.set_status("Processing…")

        threading.Thread(
            target=self._run,
            args=(import_path, output_path, invoice_date),
            daemon=True,
        ).start()

    def _run(self, import_path: str, output_path: str, invoice_date: str):
        try:
            self._process(import_path, output_path, invoice_date)
            self.set_status("Done!")
        except Exception as exc:
            self.log(f"\n[ERROR] {exc}")
            self.log(traceback.format_exc())
            self.set_status("Error — see log for details.")
        finally:
            self.progress.stop()
            self.gen_btn.state(["!disabled"])
            self.open_btn.state(["!disabled"])

    def _process(self, import_path: str, output_path: str, invoice_date: str):
        self.log(f"Import file : {os.path.basename(import_path)}")
        self.log(f"Output      : {output_path}")
        self.log(f"Invoice date: {invoice_date}")
        self.log("-" * 60)

        self.set_status("Reading import file…")
        invoice_data, bill_data = read_import(import_path)

        clients = list(invoice_data.keys())
        self.log(f"Clients found: {', '.join(clients)}\n")

        for idx, client_code in enumerate(clients, 1):
            info = invoice_data[client_code]
            transactions = bill_data.get(client_code, [])
            self.set_status(f"Generating {client_code} ({idx}/{len(clients)})…")
            self.log(f"[{idx}/{len(clients)}] {client_code} — {len(transactions)} transaction(s)")

            # Use date from import file if provided, otherwise use UI date
            imported_date = info.get("開立日期")
            if imported_date:
                if hasattr(imported_date, "strftime"):
                    eff_date = imported_date.strftime("%Y/%m/%d")
                else:
                    eff_date = str(imported_date).split(" ")[0].replace("-", "/")
            else:
                eff_date = invoice_date

            generate_client_files(
                client_code=client_code,
                info=info,
                transactions=transactions,
                output_dir=output_path,
                invoice_date=eff_date,
                log_fn=self.log,
            )

        self.log("\n" + "=" * 60)
        self.log(f"All done!  {len(clients)} client(s) processed.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
